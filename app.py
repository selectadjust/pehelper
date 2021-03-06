from flask import Flask, render_template, request, send_file
import openpyxl

app = Flask(__name__, static_folder='static', static_url_path='')

#main loading code
@app.route('/')
def entering():
    kosha_listdata = open('x.txt', 'r', encoding='utf-8')
    koshalist = kosha_listdata.readlines()
    guideList = []
    for view_list in koshalist:
        guideList.append(view_list.replace('\n',''))

    return render_template('home.html', guidelists=guideList)

# Search code
@app.route('/find', methods=['POST'])
def finding(keyword=None):
    if request.method == 'POST':
        finding_word = request.form['keyword']
    else:
        finding_word = None

    x = open('keyword recoder.txt', 'a', encoding='utf-8')
    x.write(finding_word)
    x.close()
    
    data_base = openpyxl.load_workbook('./Guide_DB.xlsx')
    data_base = data_base.get_sheet_by_name('Sheet1')

    guide_name = []
    guide_page = []
    guide_contents = []

    #empty list add guide contents
    for contents in data_base.rows:
        guide_name.append(str(contents[0].value))
        guide_page.append(str(contents[1].value))
        guide_contents.append(str(contents[2].value))

    finding_name = []
    finding_page = []
    finding_contents = []

    #finding contents added to temp database for detail seraching
    count = 0    
    for searchWord in guide_contents:
        count = count + 1
        if finding_word in searchWord:
            finding_contents.append(searchWord)
            finding_name.append(guide_name[count-1])
            finding_page.append(guide_page[count-1])
    
    temp_db = openpyxl.load_workbook('./Temp_DB.xlsx')
    temp_db_ws1 = temp_db.worksheets[0]
    temp_db.remove(temp_db_ws1)
    temp_db_ws = temp_db.create_sheet('find')

    numberOfcount = len(finding_contents)

    rewrite_count1 = 0
    rewrite_count2 = 0
    rewrite_count3 = 0

    for result_name in finding_name:
        rewrite_count1 = rewrite_count1 + 1
        temp_db_ws.cell(row=rewrite_count1, column=1).value = result_name

    for result_page in finding_page:
        rewrite_count2 = rewrite_count2 + 1
        temp_db_ws.cell(row=rewrite_count2, column=2).value = result_page

    for result_contents in finding_contents:
        rewrite_count3 = rewrite_count3 + 1
        temp_db_ws.cell(row=rewrite_count3, column=3).value = result_contents

    temp_db.save('./Temp_DB.xlsx')

    link_db = openpyxl.load_workbook('./link.xlsx')
    link_db = link_db.get_sheet_by_name('Sheet1')

    link_name = []
    link_path = []

    for link_write in link_db.rows:
        link_name.append(str(link_write[0].value))
        link_path.append(str(link_write[1].value))
    
     

    link_data = {
        
    }

    link_cont = 0

    for link_dic in link_name:
        link_data[link_name[link_cont]] = link_path[link_cont]
        link_cont = link_cont + 1

    link_link = []

    for linkserach in finding_name:
        link_link.append(link_data.get(linkserach))


    return render_template(
        'home_find.html', 
        test=finding_name, 
        test2=finding_page, 
        test3=finding_contents, 
        findcount=numberOfcount, 
        adress=link_link)

@app.route('/detailfind', methods=['POST'])
def detailsearching(detail_keyword=None):
    if request.method == 'POST':
        detail_keyword = request.form['detail_keyword']
    else:
        detail_keyword = None

    searching_list = openpyxl.load_workbook('./Temp_DB.xlsx')
    searching_sheet = searching_list.worksheets[0]
    
    detail_name =[]
    detail_page =[]
    detail_contents = []

    for detail in searching_sheet.rows:
        detail_name.append(str(detail[0].value))
        detail_page.append(str(detail[1].value))
        detail_contents.append(str(detail[2].value))

    searching_list.remove_sheet(searching_sheet)

    detail_find_name = []
    detail_find_page = []
    detail_find_contents = []

    countdetail = 0    
    for detailsearching in detail_contents:
        countdetail = countdetail + 1
        if detail_keyword in detailsearching:
            detail_find_contents.append(detailsearching)
            detail_find_name.append(detail_name[countdetail-1])
            detail_find_page.append(detail_page[countdetail-1])

    searching_newSheet = searching_list.create_sheet('find')

    name_count = 0
    page_count = 0
    contents_count = 0

    for detail_name1 in detail_find_name:
        name_count = name_count + 1
        searching_newSheet.cell(row=name_count, column=1).value = detail_name1

    for detail_page1 in detail_find_page:
        page_count = page_count + 1
        searching_newSheet.cell(row=page_count, column=2).value = detail_page1

    for detail_contents1 in detail_find_contents:
        contents_count = contents_count + 1
        searching_newSheet.cell(row=contents_count, column=3).value = detail_contents1

    searching_list.save('./Temp_DB.xlsx')

    link_db = openpyxl.load_workbook('./link.xlsx')
    link_db = link_db.get_sheet_by_name('Sheet1')

    link_name = []
    link_path = []

    for link_write in link_db.rows:
        link_name.append(str(link_write[0].value))
        link_path.append(str(link_write[1].value))
    
     

    link_data = {
        
    }

    link_cont = 0

    for link_dic in link_name:
        link_data[link_name[link_cont]] = link_path[link_cont]
        link_cont = link_cont + 1

    link_link = []

    for linkserach in detail_find_name:
        link_link.append(link_data.get(linkserach))

    return render_template(
        'home_find.html',
         test=detail_find_name, 
         test2=detail_find_page, 
         test3=detail_find_contents, 
         adress=link_link)

#kosha guide downlaod code
@app.route('/download/<path:file>')
def download(file):
    guide_n = file
    guide_a = 'templates/Guide/'+guide_n
    return send_file(guide_a)

#favicon loaidng code
@app.route('/favicon.ico')
def favicon():
    return send_file('favicon.ico')

#robots loading code
@app.route('/robots.txt')
def robots():
    return send_file('robots.txt')

    
if __name__ == '__main__':
    app.run(debug='ture')